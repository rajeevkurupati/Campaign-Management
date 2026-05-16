#!/usr/bin/env python3
"""
eval_export.py — Run both AI agent evals and export full results to Excel.

Sheets produced:
  1. Extractor Cases   — one row per case, per-check columns, colour-coded
  2. Validator Cases   — one row per case, gate/score/issues detail
  3. Issues Detail     — one row per issue raised by the validator
  4. Summary           — aggregate metrics for both agents

Usage:
    python eval_export.py
    python eval_export.py --out my_eval_results.xlsx
"""
import sys, argparse, time, os
sys.path.insert(0, ".")
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ── Palette ───────────────────────────────────────────────────────────────────
PASS_FILL   = PatternFill("solid", fgColor="D1FAE5")   # green-100
FAIL_FILL   = PatternFill("solid", fgColor="FEE2E2")   # red-100
WARN_FILL   = PatternFill("solid", fgColor="FEF3C7")   # amber-100
NA_FILL     = PatternFill("solid", fgColor="F3F4F6")   # gray-100
HDR_FILL    = PatternFill("solid", fgColor="4C1D95")   # deep purple
HDR2_FILL   = PatternFill("solid", fgColor="6D28D9")   # med purple
SUBHDR_FILL = PatternFill("solid", fgColor="EDE9FE")   # light purple
ROW_ALT     = PatternFill("solid", fgColor="FAFAFA")
CRIT_FILL   = PatternFill("solid", fgColor="DC2626")   # red for false-neg

HDR_FONT    = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
HDR2_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
SUBHDR_FONT = Font(name="Calibri", bold=True, color="4C1D95", size=9)
BODY_FONT   = Font(name="Calibri", size=9)
BOLD_FONT   = Font(name="Calibri", bold=True, size=9)
CRIT_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
WRAP_ALIGN  = Alignment(wrap_text=True, vertical="top")
CTR_ALIGN   = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP_ALIGN   = Alignment(vertical="top")

def thin_border():
    side = Side(border_style="thin", color="E5E7EB")
    return Border(left=side, right=side, top=side, bottom=side)

def _cell(ws, row, col, value, fill=None, font=None, alignment=None, border=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:       c.fill       = fill
    if font:       c.font       = font
    if alignment:  c.alignment  = alignment
    if border:     c.border     = border
    if number_format: c.number_format = number_format
    return c

def _hdr_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        c.fill      = HDR_FILL
        c.font      = HDR_FONT
        c.alignment = CTR_ALIGN
        c.border    = thin_border()

def _result_badge(ws, row, col, passed, na=False):
    if na:
        _cell(ws, row, col, "N/A",  NA_FILL,   BODY_FONT, CTR_ALIGN, thin_border())
    elif passed:
        _cell(ws, row, col, "PASS", PASS_FILL, Font(name="Calibri", bold=True, color="065F46", size=9), CTR_ALIGN, thin_border())
    else:
        _cell(ws, row, col, "FAIL", FAIL_FILL, Font(name="Calibri", bold=True, color="991B1B", size=9), CTR_ALIGN, thin_border())

def _set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ══════════════════════════════════════════════════════════════════════════════
# RUN EVALS — collect structured results
# ══════════════════════════════════════════════════════════════════════════════

def run_all_evals():
    from eval_suite import (EXTRACTOR_CASES, VALIDATOR_CASES,
                             _eval_extractor_case, _eval_validator_case)

    print("Running Extractor eval …")
    ext_results = []
    for i, case in enumerate(EXTRACTOR_CASES, 1):
        print(f"  EX-{i:02d}/{len(EXTRACTOR_CASES)} {case['label']}")
        res = _eval_extractor_case(case)
        ext_results.append((case, res))

    print("Running Validator eval …")
    val_results = []
    for i, case in enumerate(VALIDATOR_CASES, 1):
        print(f"  VAL-{i:02d}/{len(VALIDATOR_CASES)} {case['label']}")
        res = _eval_validator_case(case)
        val_results.append((case, res))

    return ext_results, val_results


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — EXTRACTOR CASES
# ══════════════════════════════════════════════════════════════════════════════

EXT_CHECKS = [
    "completeness", "enum_compliance", "date_validity",
    "campaign_type", "objective", "target_number",
    "locations_recall", "segments_recall", "tone",
    "start_date", "end_date",
]

def build_extractor_sheet(ws, ext_results):
    ws.title = "Extractor Cases"
    ws.freeze_panes = "D3"

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:R1")
    c = ws.cell(row=1, column=1,
                value="Extractor Agent — Golden Dataset Results")
    c.fill      = HDR_FILL
    c.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    c.alignment = CTR_ALIGN
    ws.row_dimensions[1].height = 22

    # ── Section headers ───────────────────────────────────────────────────────
    # Column A–C: case info
    # Column D–N: per-check results  (11 checks)
    # Column O–R: tokens / timing / overall

    info_hdrs  = ["Case ID", "Label", "Brief"]
    check_hdrs = [c.replace("_", " ").title() for c in EXT_CHECKS]
    meta_hdrs  = ["Overall", "Tokens In", "Tokens Out", "Time (s)"]

    # Group header row 2
    group_spans = [
        (1,  3,  "Case Info",      HDR2_FILL),
        (4,  14, "Per-Check Results", PatternFill("solid", fgColor="065F46")),
        (15, 18, "Metrics",        PatternFill("solid", fgColor="1D4ED8")),
    ]
    for start, end, label, fill in group_spans:
        ws.merge_cells(start_row=2, start_column=start,
                       end_row=2,   end_column=end)
        c = ws.cell(row=2, column=start, value=label)
        c.fill = fill; c.font = HDR_FONT; c.alignment = CTR_ALIGN
    ws.row_dimensions[2].height = 18

    # Column header row 3
    all_hdrs = info_hdrs + check_hdrs + meta_hdrs
    _hdr_row(ws, 3, all_hdrs)
    ws.row_dimensions[3].height = 32

    # ── Data rows ─────────────────────────────────────────────────────────────
    for r_idx, (case, res) in enumerate(ext_results):
        row = 4 + r_idx
        fill = ROW_ALT if r_idx % 2 else PatternFill("solid", fgColor="FFFFFF")
        ws.row_dimensions[row].height = 52

        checks   = res.get("checks", {})
        usage    = res.get("usage", {})
        overall  = res.get("ok", False) and all(v["pass"] for v in checks.values())

        # Case info
        for col, val in enumerate([
            case["id"],
            case["label"],
            case["brief"][:180],
        ], 1):
            _cell(ws, row, col, val, fill, BODY_FONT, WRAP_ALIGN, thin_border())

        # Per-check columns
        for c_idx, ck in enumerate(EXT_CHECKS, 4):
            if ck in checks:
                cv = checks[ck]
                badge_fill = PASS_FILL if cv["pass"] else FAIL_FILL
                badge_font = Font(name="Calibri", size=9,
                                  bold=True,
                                  color="065F46" if cv["pass"] else "991B1B")
                detail = ("✓ " if cv["pass"] else "✗ ") + cv["detail"]
                _cell(ws, row, c_idx, detail, badge_fill, badge_font, WRAP_ALIGN, thin_border())
            else:
                _cell(ws, row, c_idx, "—", NA_FILL, BODY_FONT, CTR_ALIGN, thin_border())

        # Overall
        _result_badge(ws, row, 15, overall)

        # Tokens / timing
        for col, val in enumerate([
            usage.get("input_tokens",  "—"),
            usage.get("output_tokens", "—"),
            res.get("elapsed", "—"),
        ], 16):
            _cell(ws, row, col, val, fill, BODY_FONT, CTR_ALIGN, thin_border())

        # Error row
        if not res.get("ok"):
            ws.merge_cells(start_row=row, start_column=4,
                           end_row=row,   end_column=14)
            _cell(ws, row, 4,
                  "ERROR: " + res.get("error", "unknown"),
                  FAIL_FILL, Font(name="Calibri", bold=True, color="991B1B", size=9),
                  WRAP_ALIGN, thin_border())

    _set_col_widths(ws, [9, 28, 42] + [22]*11 + [9, 10, 10, 8])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — VALIDATOR CASES
# ══════════════════════════════════════════════════════════════════════════════

def build_validator_sheet(ws, val_results):
    ws.title = "Validator Cases"
    ws.freeze_panes = "E3"

    ws.merge_cells("A1:V1")
    c = ws.cell(row=1, column=1,
                value="Content Validator Agent — Golden Dataset Results")
    c.fill      = HDR_FILL
    c.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    c.alignment = CTR_ALIGN
    ws.row_dimensions[1].height = 22

    # Group headers (row 2)
    groups = [
        (1,  4,  "Case Info",           HDR2_FILL),
        (5,  8,  "Campaign Content",    PatternFill("solid", fgColor="1D4ED8")),
        (9,  14, "Gate & Score",        PatternFill("solid", fgColor="065F46")),
        (15, 18, "Issues",              PatternFill("solid", fgColor="D97706")),
        (19, 22, "Metrics",             PatternFill("solid", fgColor="374151")),
    ]
    for start, end, label, fill in groups:
        ws.merge_cells(start_row=2, start_column=start,
                       end_row=2,   end_column=end)
        c = ws.cell(row=2, column=start, value=label)
        c.fill = fill; c.font = HDR_FONT; c.alignment = CTR_ALIGN
    ws.row_dimensions[2].height = 18

    # Column headers (row 3)
    hdrs = [
        # Case info (A–D)
        "Case ID", "Label", "Category", "Brief",
        # Campaign content (E–H)
        "Headline", "Body Copy", "Campaign Type", "Brief (full)",
        # Gate & score (I–N)
        "Expected Gate", "Actual Gate", "Gate Match",
        "Expected Score", "Actual Score", "Score In Range",
        # Issues (O–R)
        "False Negative ⚠", "False Positive", "Issue Count", "Issues Detail",
        # Metrics (S–V)
        "Summary", "Tokens In", "Tokens Out", "Time (s)",
    ]
    _hdr_row(ws, 3, hdrs)
    ws.row_dimensions[3].height = 36

    cat_fills = {
        "clear_pass":  PatternFill("solid", fgColor="D1FAE5"),
        "clear_block": PatternFill("solid", fgColor="FEE2E2"),
        "borderline":  PatternFill("solid", fgColor="FEF3C7"),
    }

    for r_idx, (case, res) in enumerate(val_results):
        row = 4 + r_idx
        base_fill = ROW_ALT if r_idx % 2 else PatternFill("solid", fgColor="FFFFFF")
        ws.row_dimensions[row].height = 72

        cat    = case["category"]
        usage  = res.get("usage", {})
        issues = res.get("issues", [])

        kw_found   = res.get("kw_found", {})
        kw_ok_str  = ", ".join(k for k, v in kw_found.items() if v)
        kw_bad_str = ", ".join(k for k, v in kw_found.items() if not v)

        issue_detail = "\n".join(
            f"[{iss.severity.upper()}] {iss.field}: {iss.message}"
            for iss in issues
        )

        data = [
            # Case info
            case["id"],
            case["label"],
            cat.replace("_", " ").upper(),
            case["brief"][:160],
            # Content
            case.get("headline", ""),
            (case.get("body_copy") or "")[:120],
            case.get("campaign_type", ""),
            case["brief"],
            # Gate & score
            case["expected_gate"].upper(),
            res["gate"].upper(),
            "PASS" if res["gate_match"] else "FAIL",
            case["expected_score_min"],
            res["score"],
            "PASS" if res["score_in_range"] else "OUT OF RANGE",
            # Issues
            "YES ⚠" if res["false_negative"] else "No",
            "Yes" if res["false_positive"] else "No",
            len(issues),
            issue_detail,
            # Metrics
            res.get("summary", ""),
            usage.get("input_tokens",  ""),
            usage.get("output_tokens", ""),
            res.get("elapsed", ""),
        ]

        for col, val in enumerate(data, 1):
            cell_fill = base_fill

            # Category badge (col 3)
            if col == 3:
                cell_fill = cat_fills.get(cat, base_fill)

            # Gate match (col 11)
            elif col == 11:
                cell_fill = PASS_FILL if res["gate_match"] else FAIL_FILL
                _cell(ws, row, col, val, cell_fill,
                      Font(name="Calibri", bold=True, size=9,
                           color="065F46" if res["gate_match"] else "991B1B"),
                      CTR_ALIGN, thin_border())
                continue

            # Score in range (col 14)
            elif col == 14:
                cell_fill = PASS_FILL if res["score_in_range"] else WARN_FILL
                _cell(ws, row, col, val, cell_fill,
                      Font(name="Calibri", bold=True, size=9,
                           color="065F46" if res["score_in_range"] else "92400E"),
                      CTR_ALIGN, thin_border())
                continue

            # False negative (col 15) — critical
            elif col == 15:
                if res["false_negative"]:
                    _cell(ws, row, col, val, CRIT_FILL, CRIT_FONT, CTR_ALIGN, thin_border())
                else:
                    _cell(ws, row, col, "No", PASS_FILL,
                          Font(name="Calibri", bold=True, color="065F46", size=9),
                          CTR_ALIGN, thin_border())
                continue

            # False positive (col 16)
            elif col == 16:
                cell_fill = WARN_FILL if res["false_positive"] else PASS_FILL
                _cell(ws, row, col, val, cell_fill, BODY_FONT, CTR_ALIGN, thin_border())
                continue

            # Actual score (col 13) — number format
            elif col == 13:
                c_obj = _cell(ws, row, col, val, cell_fill, BOLD_FONT, CTR_ALIGN, thin_border())
                continue

            _cell(ws, row, col, val, cell_fill, BODY_FONT, WRAP_ALIGN, thin_border())

    # Keyword info appended after last data row
    kw_row = 4 + len(val_results) + 1
    ws.merge_cells(start_row=kw_row, start_column=1, end_row=kw_row, end_column=22)
    c = ws.cell(row=kw_row, column=1,
                value="Note: 'Issues Detail' contains verbatim LLM issue messages. "
                      "'False Negative' = violating content that received a PASS gate — "
                      "the most critical failure mode.")
    c.fill = SUBHDR_FILL; c.font = SUBHDR_FONT; c.alignment = WRAP_ALIGN
    ws.row_dimensions[kw_row].height = 30

    _set_col_widths(ws, [
        9, 26, 14, 36,          # case info
        30, 36, 14, 36,          # content
        13, 13, 11, 14, 11, 13, # gate/score
        14, 13, 10, 55,          # issues
        48, 10, 10, 9,           # metrics
    ])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — ISSUES DETAIL  (one row per issue raised)
# ══════════════════════════════════════════════════════════════════════════════

def build_issues_sheet(ws, val_results):
    ws.title = "Validator Issues Detail"
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1,
                value="Validator — All Issues Raised by the LLM (one row per issue)")
    c.fill = HDR_FILL; c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    c.alignment = CTR_ALIGN
    ws.row_dimensions[1].height = 22

    hdrs = ["Case ID", "Label", "Category", "Gate", "Score",
            "Field", "Severity", "Issue Message"]
    _hdr_row(ws, 2, hdrs)
    ws.row_dimensions[2].height = 30

    row = 3
    sev_fills = {
        "error":   PatternFill("solid", fgColor="FEE2E2"),
        "warning": PatternFill("solid", fgColor="FEF3C7"),
    }
    sev_fonts = {
        "error":   Font(name="Calibri", bold=True, color="991B1B", size=9),
        "warning": Font(name="Calibri", bold=True, color="92400E", size=9),
    }

    for r_idx, (case, res) in enumerate(val_results):
        if not res.get("issues"):
            continue
        for iss in res["issues"]:
            fill = sev_fills.get(iss.severity, ROW_ALT)
            data = [
                case["id"],
                case["label"],
                case["category"].replace("_"," ").upper(),
                res["gate"].upper(),
                res["score"],
                iss.field,
                iss.severity.upper(),
                iss.message,
            ]
            for col, val in enumerate(data, 1):
                cell_fill = (fill if col >= 6 else
                             (ROW_ALT if r_idx % 2 else PatternFill("solid", fgColor="FFFFFF")))
                font = (sev_fonts.get(iss.severity, BODY_FONT) if col == 7
                        else BODY_FONT)
                _cell(ws, row, col, val, cell_fill, font, WRAP_ALIGN, thin_border())
            ws.row_dimensions[row].height = 50
            row += 1

    _set_col_widths(ws, [9, 26, 14, 9, 7, 13, 10, 80])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

def build_summary_sheet(ws, ext_results, val_results):
    ws.title = "Summary"

    # Title
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value="HerKey AI Agent Eval Suite — Summary")
    c.fill = HDR_FILL; c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    c.alignment = CTR_ALIGN
    ws.row_dimensions[1].height = 26

    row = 3

    # ── EXTRACTOR section ────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row=row, column=1, value="EXTRACTOR AGENT")
    c.fill = PatternFill("solid", fgColor="065F46")
    c.font = HDR_FONT; c.alignment = CTR_ALIGN
    ws.row_dimensions[row].height = 20
    row += 1

    # Per-check aggregates
    _hdr_row(ws, row, ["Check", "Passed", "Total", "Pass Rate", "Status", "Notes"], col_start=1)
    ws.row_dimensions[row].height = 20
    row += 1

    # Collect all checks across all results
    check_map = {}
    for case, res in ext_results:
        for ck, cv in res.get("checks", {}).items():
            check_map.setdefault(ck, []).append(cv["pass"])

    total_p, total_n = 0, 0
    for ck in sorted(check_map):
        passes = sum(check_map[ck])
        total  = len(check_map[ck])
        pct    = 100 * passes / total if total else 0
        status = "PASS" if pct >= 90 else ("WARN" if pct >= 70 else "FAIL")
        fill   = PASS_FILL if pct >= 90 else (WARN_FILL if pct >= 70 else FAIL_FILL)
        notes  = ""
        if ck == "date_validity" and pct < 100:
            notes = "Single-day event brief → start=end (strict check)"
        if ck == "objective" and pct < 100:
            notes = "Ambiguous brief → AI chose Community Growth vs Registrations"

        for col, val in enumerate(
            [ck.replace("_"," ").title(), passes, total, f"{pct:.0f}%", status, notes], 1
        ):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill if col in (4,5) else (ROW_ALT if (row % 2 == 0) else PatternFill("solid", fgColor="FFFFFF"))
            c.font = BODY_FONT; c.alignment = WRAP_ALIGN; c.border = thin_border()
        total_p += passes; total_n += total
        row += 1

    # Overall extractor row
    pct = 100 * total_p / total_n if total_n else 0
    fill = PASS_FILL if pct >= 90 else WARN_FILL
    for col, val in enumerate(["OVERALL", total_p, total_n, f"{pct:.1f}%", "PASS" if pct >= 90 else "FAIL", ""], 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill; c.font = Font(name="Calibri", bold=True, size=9); c.border = thin_border(); c.alignment = CTR_ALIGN
    row += 2

    # ── VALIDATOR section ────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row=row, column=1, value="CONTENT VALIDATOR AGENT")
    c.fill = PatternFill("solid", fgColor="D97706")
    c.font = HDR_FONT; c.alignment = CTR_ALIGN
    ws.row_dimensions[row].height = 20
    row += 1

    _hdr_row(ws, row, ["Metric", "Value", "Target", "Status", "Critical?", "Notes"], col_start=1)
    ws.row_dimensions[row].height = 20
    row += 1

    cats = {"clear_pass": [], "clear_block": [], "borderline": []}
    fn_cases, fp_cases = [], []
    for case, res in val_results:
        cats[case["category"]].append(res)
        if res["false_negative"]: fn_cases.append((case, res))
        if res["false_positive"]: fp_cases.append((case, res))

    total_gm = sum(r["gate_match"]  for _, r in val_results)
    total_sr = sum(r["score_in_range"] for _, r in val_results)
    total_n  = len(val_results)

    val_metrics = [
        ("Overall Gate Accuracy",     f"{100*total_gm/total_n:.0f}%",  "≥ 90%",  total_gm/total_n >= 0.9,    False,
         "Correct PASS/BLOCK decision"),
        ("Overall Score Range Acc.",  f"{100*total_sr/total_n:.0f}%",  "≥ 80%",  total_sr/total_n >= 0.8,    False,
         "Score within expected range"),
        ("Clear-PASS gate accuracy",  f"{100*sum(r['gate_match'] for r in cats['clear_pass'])/max(len(cats['clear_pass']),1):.0f}%",
         "100%", all(r["gate_match"] for r in cats["clear_pass"]), False, "Compliant content not blocked"),
        ("Clear-BLOCK gate accuracy", f"{100*sum(r['gate_match'] for r in cats['clear_block'])/max(len(cats['clear_block']),1):.0f}%",
         "100%", all(r["gate_match"] for r in cats["clear_block"]), True, "Violating content always blocked"),
        ("False Negative count",      len(fn_cases),                   "0",       len(fn_cases) == 0,          True,
         "Violating content that passed — most critical"),
        ("False Positive count",      len(fp_cases),                   "≤ 1",     len(fp_cases) <= 1,          False,
         "Compliant content that was blocked"),
        ("Borderline gate accuracy",  f"{100*sum(r['gate_match'] for r in cats['borderline'])/max(len(cats['borderline']),1):.0f}%",
         "≥ 70%", sum(r["gate_match"] for r in cats["borderline"])/max(len(cats["borderline"]),1) >= 0.7,
         False, "Score near threshold — inherently uncertain"),
    ]

    for metric, val, target, ok, critical, notes in val_metrics:
        status = "PASS" if ok else "FAIL"
        fill   = PASS_FILL if ok else (CRIT_FILL if critical else FAIL_FILL)
        s_font = (CRIT_FONT if (not ok and critical)
                  else Font(name="Calibri", bold=True, color="065F46" if ok else "991B1B", size=9))
        base_fill = ROW_ALT if (row % 2 == 0) else PatternFill("solid", fgColor="FFFFFF")

        for col, cval in enumerate([metric, val, target, status, "YES" if critical else "—", notes], 1):
            c = ws.cell(row=row, column=col, value=cval)
            c.font   = s_font if col == 4 else (Font(name="Calibri", bold=True, size=9) if col == 5 and critical else BODY_FONT)
            c.fill   = fill   if col == 4 else base_fill
            c.border = thin_border()
            c.alignment = CTR_ALIGN if col in (2,3,4,5) else WRAP_ALIGN
        row += 1

    # False negative detail
    if fn_cases:
        row += 1
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value="FALSE NEGATIVE DETAIL — review and tighten system prompt")
        c.fill = CRIT_FILL; c.font = CRIT_FONT; c.alignment = WRAP_ALIGN
        row += 1
        for case, res in fn_cases:
            ws.merge_cells(f"A{row}:F{row}")
            c = ws.cell(row=row, column=1,
                value=f"{case['id']}  |  {case['label']}  |  score={res['score']}  |  "
                      f"summary: {res.get('summary','')}")
            c.fill = FAIL_FILL; c.font = BODY_FONT; c.alignment = WRAP_ALIGN
            ws.row_dimensions[row].height = 28
            row += 1

    # ── Token cost summary ────────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    c = ws.cell(row=row, column=1, value="TOKEN USAGE SUMMARY")
    c.fill = HDR2_FILL; c.font = HDR_FONT; c.alignment = CTR_ALIGN
    row += 1

    _hdr_row(ws, row, ["Agent", "Cases", "Avg Tokens In", "Avg Tokens Out", "Avg Time (s)", "Est. Cost note"])
    row += 1

    import statistics as _st
    for agent_label, results in [("Extractor", ext_results), ("Validator", val_results)]:
        usages = [r.get("usage",{}) for _,r in results if r.get("usage")]
        times  = [r.get("elapsed",0) for _,r in results if r.get("elapsed")]
        if usages:
            avg_in  = _st.mean(u.get("input_tokens",0)  for u in usages)
            avg_out = _st.mean(u.get("output_tokens",0) for u in usages)
            avg_t   = _st.mean(times) if times else 0
            note = "Cache hits reduce cost for repeated validations"
        else:
            avg_in = avg_out = avg_t = 0
            note = "No usage data"
        data = [agent_label, len(results), f"{avg_in:.0f}", f"{avg_out:.0f}", f"{avg_t:.1f}s", note]
        for col, val in enumerate(data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = ROW_ALT; c.font = BODY_FONT; c.alignment = CTR_ALIGN; c.border = thin_border()
        row += 1

    _set_col_widths(ws, [34, 14, 10, 10, 10, 50])


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default="eval_results.xlsx",
                        help="Output Excel filename (default: eval_results.xlsx)")
    args = parser.parse_args()

    out_path = args.out
    if not os.path.isabs(out_path):
        out_path = os.path.join(os.path.dirname(__file__), out_path)

    t0 = time.time()
    ext_results, val_results = run_all_evals()

    print("\nBuilding workbook …")
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    ws_summary   = wb.create_sheet("Summary")
    ws_ext       = wb.create_sheet("Extractor Cases")
    ws_val       = wb.create_sheet("Validator Cases")
    ws_issues    = wb.create_sheet("Validator Issues Detail")

    build_summary_sheet(ws_summary,   ext_results, val_results)
    build_extractor_sheet(ws_ext,     ext_results)
    build_validator_sheet(ws_val,     val_results)
    build_issues_sheet(ws_issues,     val_results)

    wb.save(out_path)
    elapsed = time.time() - t0
    print(f"\nSaved: {out_path}")
    print(f"Sheets: Summary | Extractor Cases | Validator Cases | Validator Issues Detail")
    print(f"Total time: {elapsed:.1f}s")


if __name__ == "__main__":
    main()
